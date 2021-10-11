# Author: Shivansh Shalabh

# Importing module (Installation required - Selenium, Openpyxl, Pynput)
from selenium import webdriver
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from pynput.keyboard import Key, Controller
from time import sleep
import sys
import os
from datetime import datetime, date
import re

# Function to check if Hexadecimal code is valid


def isValidHexaCode(str):
    # Regex to check valid
    # hexadecimal color code.
    regex = "^#([A-Fa-f0-9]{6}|[A-Fa-f0-9]{3})$"

    # Compile the ReGex
    p = re.compile(regex)

    # If the string is empty
    # return false
    if(str == None):
        return False

    # Return if the string
    # matched the ReGex
    if(re.search(p, str)):
        return True
    else:
        return False


# Function to extract names form excel file


def get_name_xl(temp_file_name, column_skip=1, row_skip=1):
    name_list_lst = os.listdir('Name List')
    name_lst_file_name = ""
    to_remove = []
    for i in name_list_lst:
        if i.endswith('.xlsx') == True or i.endswith('.xlsm') == True or i.endswith('.xltx') == True or i.endswith('.xltm') == True:
            pass
        else:
            to_remove.append(i)
    for i in to_remove:
        name_list_lst.remove(i)
    get_only_name = False
    if not name_list_lst:
        print("No Excel file was found :(")
        print(
            "Enter 1 to get just the name of all participants.", "Enter 2 to exit", sep="\n")
        choice = input("Enter your choice: ")
        while choice not in ['1', '2', '3']:
            choice = input(
                "Enter 1 to get just the name of all participants.", "Enter 2 to exit", sep="\n")
        if choice == '1':
            print("The names of participants will be saved in a txt file with the name as",
                  temp_file_name, "in the current working directory.")
            return 0
        elif choice == '2':
            print("Thank you for using Zoomobot")
            sys.exit()
    if len(name_list_lst) == 1:
        name_lst_file_name = name_list_lst[0]
    else:
        print("Choose the file: ")
        for i in name_list_lst:
            print('\t', name_list_lst.index(i)+1, 'for', i)

        file_index = int(input("Your choice: "))
        while not file_index and file_index.isdigit() and int(file_index) in range(0, len(name_list_lst)):
            file_index = int(input())
        name_lst_file_name = name_list_lst[file_index-1]
    lst_str = []
    # Using openpyxl to open excel file and read names
    wb = load_workbook(filename='Name List/'+name_lst_file_name)
    ws = wb.active
    for i in range(row_skip, ws.max_row+1):
        if ws.cell(row=(i+1), column=(column_skip+1)).value == None:
            break
        else:
            lst_str.append(ws.cell(row=(i+1), column=(column_skip+1)).value)
    return ["\n".join(lst_str), 'Name List/'+name_lst_file_name]

# Function to mark attendance in the excel file


def mark_xl(filename, present, meeting_topic, color, row_skip=1, column_skip=1):
    for i in range(len(present)):
        present[i] = present[i].lower().replace(" ", "")
    if isValidHexaCode(color) == False:
        color = ''
    # open file using openpxl
    wb = load_workbook(filename=filename)
    ws = wb.active
    # get the empty column in the sheet
    empty_column = ws.max_column+1
    for i in range(row_skip, ws.max_row+1):
        if ws.cell(row=(i+1), column=(column_skip+1)).value == None:
            break
        else:
            if str(ws.cell(row=(i+1), column=(column_skip+1)).value).lower().replace(" ", "") in present:
                ws.cell(row=(i+1), column=(empty_column)).value = "Present"
            else:
                ws.cell(row=(i+1), column=(empty_column)).value = "Absent"
                if color:
                    ws.cell(row=(i+1), column=(empty_column)
                            ).fill = PatternFill("solid", fgColor=color.replace('#', ""))
    ws.cell(row=row_skip, column=empty_column).value = meeting_topic + \
        " | " + ".".join(reversed((str(date.today()).split("-"))))
    wb.save(filename=filename)
    return True

# Function to format all.txt file


def format_all_txt(file_name):
    txt_file = open(file_name, 'r+')
    content = txt_file.read()
    txt_file.close()
    content = content.split("\n")
    for i in range(1, len(content)):
        if content[i].find('. ') != -1:
            if((content[i].split(".")[0]).replace(" ", "")).isnumeric():
                content[i] = content[i][(content[i].find('.')+1):]
        content[i] = str(i)+". "+content[i]
        while(content[i].find('  ') != -1):
            content[i] = content[i].replace('  ', ' ')
    content = "\n".join(content)
    txt_file = open(file_name, 'w')
    content = txt_file.write(content)
    txt_file.close()
    return None


if __name__ == '__main__':
    print("🤖   Welcome to ZoomoBot   🤖")

    temp_file_name = str(datetime.now())+'.txt'
    # Getting details required to join the meeting
    joining_option = input(
        "How do you wanna join?\n\tEnter 1 to join the meeting using meeting id and password\n\tEnter 2 to join using link ")
    while joining_option not in ['1', '2']:
        joining_option = input(
            "Invalid input:(\nPls try again\n\nHow do you wanna join?\n\tEnter 1 to join the meeting using meeting id and password\n\tEnter 2 to join using link ")
    if joining_option == '2':
        meeting_link = input("Please enter the meeting link: ")
        while not meeting_link:
            meeting_link = input("Please enter the meeting link: ")
        meeting_link = meeting_link.replace(" ", "")
    elif joining_option == '1':
        meeting_id = input("Enter the meeting id: ")
        while not meeting_id:
            meeting_id = input("Enter the meeting id: ")
        meeting_id = meeting_id.replace(" ", "")
        meeting_link = "https://zoom.us/wc/join/"+meeting_id

        meeting_password = input("Enter the meeting password: ")
        while not meeting_password:
            meeting_password = input("Enter the meeting password: ")
    name = ''
    if os.path.isfile('Cache.txt'):
        with open('Cache.txt', 'r') as f:
            name = f.readline().split('|')[0]

    while not name:
        name = input("Enter name to join the meeting: ")
    nested_folder = input("Enter the meeting title: ")
    while not nested_folder:
        nested_folder = input("Enter the meeting title: ")

    print("Do you want to filter the participants? (Enter -1 to skip)", "Enter 1 to consider only participants who have raised their hand",
          "Enter 2 to consider only the participatns who are connected to audio", "Enter 3 to consider only participants who are connected to audio as well as whose hands are raised.", sep="\n")
    filter = input("You choice: ")
    while filter not in ['-1', '1', '2', '3']:
        filter = input("Invalid choice :(\nYou choice: ")
    # Fetching data from namelist
    lst_str = ""
    xl_file = ""
    get_only_name = False
    while not lst_str:
        try:
            with open('Cache.txt', 'r') as f:
                f.readline()
                first_line = f.readline()
                first_line.replace(" ", "")
                second_line = f.readline()
                second_line.replace(" ", "")
                column_skip = first_line.split("|")[0]
                color = f.readline().split("|")[0]
                color.replace(" ", "")
                cell_color = color
                while not column_skip.isdigit() or not column_skip:
                    column_skip = input(
                        "Enter the column number to skip: ")

                row_skip = first_line.split("|")[0]
                while not row_skip.isdigit() or not row_skip:
                    row_skip = input(
                        "Enter the column number to skip: ")
                while not isValidHexaCode(cell_color):
                    cell_color = input(
                        "Enter the color of the cell (Enter -1 to skip): ")
                    if cell_color == '-1':
                        cell_color = ''
                        break
                column_skip, row_skip = int(column_skip), int(row_skip)
        except:
            print("Unable to get data, please fill in the details manually.")
            column_skip = input("Enter the column number to skip: ")
            while not column_skip.isdigit():
                column_skip = input("Enter the column number to skip: ")
            column_skip = int(column_skip)
            row_skip = input("Enter the row number to skip: ")
            while not row_skip.isdigit():
                row_skip = input("Enter the row number to skip: ")
            row_skip = int(row_skip)
        xl_return = get_name_xl(temp_file_name, column_skip, row_skip)
        if xl_return == 0:
            get_name_only = True
            break
        else:
            lst_str = xl_return[0]
            xl_file = xl_return[1]

    print("Mr. ZoomoBot 👨‍💻 is joining the meeting, please ask the host to allow him in case waiting room is enabled.")
    # Joinig the meeting using selenium
    web = webdriver.Chrome()
    web.get(meeting_link)
    web.maximize_window()
    if joining_option == '2':
        keyboard = Controller()
        for _ in range(3):
            check = 0
            sleep(1)
            keyboard.press(Key.esc)
            keyboard.release(Key.esc)
            sleep(1)
            launch_btn = web.find_element_by_xpath(
                '//*[@id="zoom-ui-frame"]/div[2]/div/div[1]/div')
            launch_btn.click()
            check = 0
            sleep(1)
        sleep(1)
        keyboard.press(Key.esc)
        keyboard.release(Key.esc)
        keyboard.press(Key.esc)
        keyboard.release(Key.esc)
        join_from_browser = web.find_element_by_xpath(
            '//*[@id="zoom-ui-frame"]/div[2]/div/div[2]/h3[2]/span/a')
        join_from_browser.click()
    web.maximize_window()
    name_input = web.find_element_by_id('inputname')
    name_input.send_keys(name)
    name_join_btn = web.find_element_by_xpath('//*[@id="joinBtn"]')
    name_join_btn.click()
    if joining_option == '1':
        password_input = web.find_element_by_id('inputpasscode')
        password_input.send_keys(meeting_password)
        password_join_btn = web.find_element_by_xpath('//*[@id="joinBtn"]')
        password_join_btn.click()
    sleep(5)

    # Opening the participant panel
    # Adding class of bot_participant_name
    # Making font size smaller to make the text readable for selenium
    check = 0
    while True:
        try:
            web.execute_script(
                """
        document.getElementsByClassName('footer')[0].classList.remove("footer--hidden");
        document.getElementsByClassName("footer-button__button ax-outline")[0].click();
    """)

            break
        except:
            check += 1
            if check == 7:
                choice = input("Are you in the meeting (y/n): ")
                while not choice.lower() in ['y', 'n']:
                    choice = input("Are you in the meeting (y/n): ")
                    if choice .lower() == 'y':
                        break

            print("Joining the meeting.\nIf waiting room is enabled, please ask the host to let me in the meeting.")
            sleep(5)
    print("Thank you for your patience, Mr. ZoomoBot 😎 is now in the meeting.")
    sleep(5)
    selenium_name_lst = ""
    check = 0
    while len(selenium_name_lst) == 0:
        try:
            web.execute_script("""
        let participant_bot_list = document.getElementById('participants-ul');
        for (let i of participant_bot_list.children) {
            i.style.backgroundColor = '#66ff69 !important';
            i.children[0].children[0].children[0].setAttribute('style', 'display:none'); //profile
            i.children[0].children[0].children[1].setAttribute('style', 'font-size: 0.1px;'); //fontsize of name
            i.setAttribute('style', 'height: 0.1px;'); //li height
            i.children[0].children[1].setAttribute('style', 'display: none;'); // icons
            i.children[0].children[0].children[1].children[0].classList.add('bot_participant_name');
        }   
       """)

            selenium_name_lst = web.find_elements_by_class_name(
                'bot_participant_name')
            sleep(2)
        except:
            check += 1
            sleep(2)
            if check == 7:
                while not input("Unable to find the participant's list. Make sure the participant list is open and then input any non-empty string."):
                    pass
                break
    all_participant = []
    unknown = []
    present = []
    sleep(1)
    for name in selenium_name_lst:
        all_participant.append(name.text)
    temp_file = open(temp_file_name, 'w')
    temp_file.write("All participants:\n"+("\n".join(all_participant)))
    temp_file.close()
    if get_only_name:
        sys.exit()
    # Converting string with names to a list
    namelst = lst_str.split('\n')
    first_names = set()
    duplicate_first_names = []
    while(namelst.count('') != 0):
        namelst.remove('')
    for i in range(len(namelst)):
        namelst[i] = namelst[i].lower()
        while(namelst[i].find("  ") != -1):
            namelst[i] = namelst[i].replace("  ", " ")
        namelst[i] = namelst[i].split(" ")
        before_number_of_elems_in_set = len(first_names)
        first_names.add(namelst[i][0])
        if before_number_of_elems_in_set == len(first_names):
            duplicate_first_names.append(namelst[i][0])
    attendance_data = "Total no. of students: "+str(len(namelst))

    # Filtering participants
    if filter == '1':
        # HandRaised
        try:
            web.execute_script(
                """
        let participant_bot_list = document.getElementById('participants-ul');
        for (let i of participant_bot_list.children) {
            let svg_there = i.children[0].children[1].children[0].children[0].children[0];
            if (svg_there != undefined) {
                i.children[0].children[0].children[1].children[0].classList.add('bot_verified_success');
            }
        }
    """)
        except:
            print("Please open Participants Panel")
            sleep(2)
    elif filter == '2':
        # Connected to Audio
        try:
            web.execute_script(
                """
        let participant_bot_list = document.getElementById('participants-ul');
        for (let i of participant_bot_list.children) {
            
            let all_icons_bot = i.children[0].children[1].children;
            let check_mute_bot = false;
            for (let i of all_icons_bot) {
                let classes_child_bot = i.children[0].classList;
                for (let k of classes_child_bot) {
                    if (k.indexOf('mute') != -1) check_mute_bot = true;
                }
            }

            if (check_mute_bot === true) {
                i.children[0].children[0].children[1].children[0].classList.add('bot_verified_success');
            }
        }
    """)
        except:
            print("Please open Participants Panel")
            sleep(2)
    elif filter == '3':
        # Both
        try:
            web.execute_script(
                """
        let participant_bot_list = document.getElementById('participants-ul');
        for (let i of participant_bot_list.children) {
            
            let all_icons_bot = i.children[0].children[1].children;
            let check_mute_bot = false;
            for (let j of all_icons_bot) {
                let classes_child_bot = j.children[0].classList;
                for (let k of classes_child_bot) {
                    if (k.indexOf('mute') != -1) check_mute_bot = true;
                }
            }
        
            let svg_there = i.children[0].children[1].children[0].children[0].children[0];
            if (check_mute_bot === true && svg_there != undefined) {
                i.children[0].children[0].children[1].children[0].classList.add('bot_verified_success');
            }
        }
            """)
        except:
            print("Please open Participants Panel")
            sleep(2)
    print("Going through the participants' list 🕵")
    if filter != '-1':
        all_participant = []
        selenium_name_lst = web.find_elements_by_class_name(
            'bot_verified_success')
        for name in selenium_name_lst:
            all_participant.append(name.text)

    for i in all_participant:
        i = i.lower()
        while(i.find("  ") != -1):
            i = i.replace("  ", " ")
        temp_name_lst = i.split(" ")
        known = False
        for k in range(len(namelst)):
            check = 0
            unique = False
            for j in range(len(namelst[k])):
                if namelst[k][j] in temp_name_lst:
                    if j == 0 and namelst[k][j] not in duplicate_first_names:
                        unique = True
                        pass
                    check += 1
            if check >= 2 or unique:
                present.append(' '.join(namelst[k]))
                namelst.remove(namelst[k])
                known = True
                break

        if not known:
            unknown.append(i)
    raw_present = list(present)

    os.remove(temp_file_name)
    if xl_file:
        mark_xl(xl_file, raw_present, nested_folder,
                cell_color, column_skip, row_skip)
        print("Excel File updated successfully")
    print("Thanks you for using ZoomoBot 😊")
    web.execute_script("""
    let participant_bot_list = document.getElementById('participants-ul');
    for (let i of participant_bot_list.children) {
        i.children[0].children[0].children[0].removeAttribute('style');
        i.children[0].children[0].children[1].removeAttribute('style');
        i.removeAttribute('style');
        i.children[0].children[1].removeAttribute('style');
    }
    """)
